#!/usr/bin/python3
# -*-coding:utf-8 -*-

# Reference:**********************************************
# @Time    : 7/31/2020 7:43 PM
# @Author  : Gaopeng.Bai
# @File    : excel_operation.py
# @User    : gaope
# @Software: PyCharm
# @Description:
# Reference:**********************************************
from openpyxl import load_workbook


class excel_operator:
    def __init__(self, path="../../data_/38_GBA.xlsx", sheet= "Bestandsverzeichnis"):
        self.path = path
        self.wb = load_workbook(path)
        self.sht = self.wb[sheet]

    def write_line(self, value_list, cols=10):
        last_ = self.lastRow()
        # write and save
        for i in range(1, cols+1):
            self.sht.cell(column=i, row=last_ + 1, value=value_list[i - 1])

        self.wb.save(self.path)

    def lastRow(self):
        # last row
        a = 0
        for i, row in enumerate(self.sht.rows):
            a = i
        return a + 1


if __name__ == '__main__':
    a = excel_operator()
    list = ["1","2","3","4","5","6","7","8","9","10", ]
    a.write_line(list)